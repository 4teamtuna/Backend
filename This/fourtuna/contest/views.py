import os
import re
import datetime
from django.shortcuts import render
from rest_framework.response import Response
from .models import Contest_info, my_contest
from rest_framework.views import APIView
from .serializers import Contest_Info_serializer, My_Contest_serializer
from rest_framework.viewsets import ModelViewSet
from django.core.files import File
from openpyxl import load_workbook 
from django.core.files.images import ImageFile
from datetime import datetime

def sanitize_filename(filename):
    return re.sub(r'[/\\:*?"<>|]', "", filename).strip()

def save_image_from_crawl(image_folder, contest_info):
    image_path = os.path.join(image_folder, f"{contest_info.title}.jpg")
    print("Image path: ", image_path)  # 경로 출력
    with open(image_path, 'rb') as img_file:
        contest_info.image.save(f"{contest_info.title}.jpg", File(img_file))
        contest_info.save()

def upload_xlsx_and_images(xlsx_path, image_folder):
    workbook = load_workbook(filename=xlsx_path)
    sheet = workbook.active
    headers = [cell.value for cell in sheet[1]]

    for row in sheet.iter_rows(min_row=2):
        column_dict = {headers[i]: cell.value for i, cell in enumerate(row)}
        column_dict = {key: value for key, value in column_dict.items() if key in [f.name for f in Contest_info._meta.fields]}

        # 'date' 칼럼에 현재 날짜와 시간을 설정합니다.
        column_dict['date'] = datetime.now()

        # 이미지 파일 이름을 가져오고, 이미지 파일의 전체 경로를 구합니다.
        image_name = column_dict.get('trimmed_title')
        if image_name:
            image_name = sanitize_filename(image_name)
            image_name = image_name+".jpg"
            image_path = os.path.join(image_folder, image_name)

            # 이미지 파일이 실제로 존재하는지 확인합니다.
            if os.path.exists(image_path):
                with open(image_path, 'rb') as image_file:
                    column_dict['image'] = ImageFile(image_file, name=image_name)

        # 데이터베이스에 데이터를 저장합니다.
        contest_info = Contest_info.objects.create(**column_dict)
        contest_info.title = sanitize_filename(contest_info.title)
        save_image_from_crawl(image_folder, contest_info)
        # 데이터가 저장되었다는 메시지를 출력합니다.
        print(f'Data has been saved to the database: {contest_info}')


class Contest_Info_API(ModelViewSet):
     queryset = Contest_info.objects.all()
     serializer_class = Contest_Info_serializer

class My_Contest_API(APIView):
    def get(self, request):
        queryset = my_contest.objects.all()
        print(queryset)
        serializer = My_Contest_serializer(queryset, many = True)
        return Response(serializer.data)

